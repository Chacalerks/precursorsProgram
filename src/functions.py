from openpyxl import load_workbook
from tools import openFileExplorer, writeSummedSubstances
from substance import Substance
from db import MongoDB
from datetime import datetime
import tkinter as tk
from tkinter import messagebox
import os
import logging
# ---------------------------------------------------------------
# Database Connection Setup
# ---------------------------------------------------------------
db = MongoDB()

def load_collection():
    """Load data collection from the database."""
    global db
    db.get_month_reports_by_instance()
    db.get_month_reports()

# ---------------------------------------------------------------
# Creation Functions and insert into DB
# ---------------------------------------------------------------
def create_report_by_instance(substances, report):
    """Create and store a report by instance in the database."""
    global db
    substances_array = [substances[key].__dict__ for key in substances]
    report["substances"] = substances_array
    report["date"] = datetime.now()
    return db.insert_monthly_report_by_instance(report)

def create_monthly_report(month, year):
    """Aggregate monthly reports, process them, and create a summarized report."""
    global db
    year = int(year)
    reports = list(db.month_report_by_instance.find({"year": year, "month": month}, {"_id": 0}))
    substances = {}
    cantInstances = 0
    if reports:
        for report in reports:
            for item in report["substances"]:
                substance = Substance(**item)
                if substance.name in substances:
                    substances[substance.name].update(substance)
                else:
                    substances[substance.name] = substance
            cantInstances +=1
    else:
        return False
    report_data = {
        "month": month,
        "year": year,
        "dateCreated": datetime.now(),
        "substances": [substances[key].__dict__ for key in substances],
        "cantInstances":cantInstances
    }
    if substances:
        return report_data
    return False

# ---------------------------------------------------------------
# Query from DB
# ---------------------------------------------------------------
def get_instances_on_month_report(month, year):
    """Aggregate monthly reports, process them, and create a summarized report."""
    global db
    year = int(year)
    reports = list(db.month_report_by_instance.find({"year": year, "month": month}, {"_id": 0}))    
    if reports:
        return len(reports)
    return 0

def check_month_report(month, year):
    """Check if a lab report exists for a given month and instance."""
    global db
    report = list(db.month_report.find({"year": year, "month": month}))
    if not report == None:
        instanceReport = get_instances_on_month_report(month, year) # get how many instances haved uploaded its report on (month, year)
        if len(report) != 0:
            report = report[0] #should be just one repor by month and year            
            if report["cantInstances"] != instanceReport: #check if need update
                tempReport = create_monthly_report(month, year)
                report["dateUpdate"] = datetime.now()
                report["substances"] = tempReport["substances"]
                report["cantInstances"] = tempReport["cantInstances"]
                db.update_monthly_report(report)
            
        elif instanceReport != 0:
            report = create_monthly_report(month, year)
            if report:
                db.insert_monthly_report(report)
            
    return report if report else False


def check_labs_report_on_month(month, instance, year):
    """Check if a lab report exists for a given month and instance."""
    global db
    reports = list(db.month_report_by_instance.find({"year": year, "month": month, "instance": instance}, {"_id": 0}))
    return reports[0] if reports else False


# ---------------------------------------------------------------
# File Processing Functions
# ---------------------------------------------------------------
def process_files(switchBD, adminView):
    """Process files for reporting, either loading or creating reports."""
    substances = {}
    files_paths = openFileExplorer()
    filenames = ""
    flagResult = []
    for file_path in files_paths:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb.active
        report = {}
        substances_temp = {}
        
        for row in sheet.iter_rows(min_row=12, values_only=True):
            if "Lista" not in str(row[1]) and row[3] != "-":
                try:
                    substance = Substance(row[1], "Kg", row[6], row[7], row[9], row[11], row[12])
                    if substance.name in substances:
                        substances[substance.name].update(substance)
                    else:
                        substances[substance.name] = substance
                except Exception as e:
                    logging.error(f"Error inserting monthly report: {e}")
                    show_message("Error","No se han guardado los archivos a la base de datos de: "+file_path)
                    raise 
            
        
        if switchBD == "on":
            report.update({
                "instance": sheet.cell(row=6, column=3).value,
                "year": sheet.cell(row=4, column=3).value,
                "month": sheet.cell(row=5, column=3).value,
                "personAssigned": sheet.cell(row=7, column=3).value,
                "email": sheet.cell(row=8, column=3).value,
                "phoneNumber": sheet.cell(row=9, column=3).value
            })
            if check_labs_report_on_month(report["month"], report["instance"],  report["year"]) == False:
                if not adminView:
                    flagResult.append(create_report_by_instance(substances_temp, report))
            else:
                show_message("Error", "El reporte "+report["month"]+" "+ str(report["year"])+" "+ report["instance"] + " ya existe", "error" )
                return 0
        wb.close()
    if adminView:
        writeSummedSubstances(substances)
    else:
        result = all(flagResult)
        file_names = [os.path.basename(filepath) for filepath in files_paths]
        if result and flagResult != []:         
            # Join the file names into a single string separated by commas
            file_names_string = ", ".join(file_names)
            show_message("Perfecto","Se ha registrado correctamente los archivos: "+file_names_string)
        elif result == False:
            # get de index of false elements
            false_index = [index for index, value in enumerate(flagResult) if not value]
            false_files_names = ', '.join([file_names[index] for index in false_index])
            show_message("Error","No se han guardado los archivos a la base de datos de: "+false_files_names)

# ---------------------------------------------------------------
# User Interaction
# ---------------------------------------------------------------
def show_message(title, msg, box_type="info"):
    """Display a custom message box to the user based on the type specified."""
    if box_type == "info":
        messagebox.showinfo(title, msg)
    elif box_type == "warning":
        messagebox.showwarning(title, msg)
    elif box_type == "error":
        messagebox.showerror(title, msg)
    else:
        messagebox.showinfo(title, msg)  # Default to showinfo if no valid type provided