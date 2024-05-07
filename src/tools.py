from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl import Workbook
from tkinter import filedialog
from tkinter.filedialog import asksaveasfilename
import datetime
import os
import sys
import locale

def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller."""
    import sys, os
    if getattr(sys, 'frozen', False):
        # we are running in a |PyInstaller| bundle
        base_path = sys._MEIPASS
    else:
        # we are running in a normal Python environment
        base_path = os.path.abspath(".")
    
    full_path = os.path.join(base_path, relative_path)
    print(f'Resource path: {full_path}')  # Debugging output
    return full_path


def scaleImg(image_path):
    img = Image(image_path)
    
    # Original dimensions in cm (as per your specifications)
    original_width_cm = 5.94
    original_height_cm = 1.96

    # Convert cm to pixels (1 inch = 2.54 cm, Excel's default DPI = 96)
    pixels_per_cm = 96 / 2.54
    width_pixels = original_width_cm * pixels_per_cm
    height_pixels = original_height_cm * pixels_per_cm

    # Adjust the image size
    # Note: Excel column width units and row height points may need to be adjusted
    column_width_unit = 9.14  # default Excel column width unit in pixels
    row_height_points = 0.75  # points per pixel in row height

    # Set size to fit within a merged cell area of 'A1:C1'
    target_width_pixels = 3 * (8.5 * column_width_unit)  # Assume each column is 8.43 characters wide
    target_height_points = 51  # Set row height as previously defined in points

    scale_width = target_width_pixels / width_pixels
    scale_height = target_height_points / (height_pixels * row_height_points)

    # Apply the smaller of the two scales to maintain aspect ratio
    scale = min(scale_width, scale_height)
    
    img.width = width_pixels * scale
    img.height = height_pixels * scale
    
    return img, target_height_points, target_width_pixels, column_width_unit

def openFileExplorer():
    filetypes = (('Excel files', '*.xlsx *.xls *.xlsm'), ('All files', '*.*'))
    filepaths = filedialog.askopenfilenames(title='Open files', initialdir='.', filetypes=filetypes)
    return filepaths

def writeSummedSubstances(substances, month_year = ""):
    wb = Workbook()
    ws = wb.active

    # Load and insert the logo image
    image_path = resource_path('src/img/GASEL_R_QUIMICA.png')  # Updated to use resource_path 
    img, target_height_points, target_width_pixels, column_width_unit = scaleImg(image_path)



    ws.merge_cells('A1:C1')  # Merge cells for the logo
    ws.add_image(img, 'A1')  # Place image in the merged cell A1

    # Set row height and column widths to accommodate the image
    ws.row_dimensions[1].height = target_height_points
    ws.column_dimensions['A'].width = target_width_pixels / 3 / column_width_unit
    ws.column_dimensions['B'].width = target_width_pixels / 3 / column_width_unit
    ws.column_dimensions['C'].width = target_width_pixels / 3 / column_width_unit
    
    
    
    
    
    ws.merge_cells('D1:F1')  # Merge for the title "Reporte de Precursores"
    ws['D1'] = 'Reporte de Precursores'
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')

    ws['G1'] = 'Código: F-06-RQ-TEC'
    ws['G1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2')  # Merge for the second row to create space

    # Column headers
    columns = ["Sustancia", "Unidad", "Ingreso Período", "# Solicitud de Bienes", "Despacho Período", "Existencias Final Período", "Uso dado a la Sustancia"]
    ws.append(columns)

    # Setting style for the header row
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue background
    font_white = Font(color="FFFFFF")  # White font

    for cell in ws[3]:  # Third row has the headers
        cell.fill = header_fill
        cell.font = font_white
    ws.row_dimensions[3].height = 45  # Third row height

    # Adjusting widths for the columns
    width_name = 30
    width_use = 60
    count = 0
    for col in ws.columns:
        if count == 0:
            ws.column_dimensions[col[2].column_letter].width = width_name
        elif count == 6:
            ws.column_dimensions[col[2].column_letter].width = width_use
        else:            
            max_length = max(len(str(cell.value)) for cell in col)
            adjusted_width = max_length + 2  # Adding some extra padding
            ws.column_dimensions[col[2].column_letter].width = adjusted_width
        count += 1

    # Iterate over substances (whether it's a list or a dict)
    items = substances.values() if isinstance(substances, dict) else substances
    for substance in items:
        if isinstance(substance, dict):
            row = [
                substance['name'], substance['unityMeasure'], substance['inventoryAdditions'],
                substance['additionNumber'], substance['inventoryExpenses'], substance['finalBalance'],
                substance['expenseJustification']
            ]
        else:
            row = [
                substance.name, substance.unityMeasure, substance.inventoryAdditions,
                substance.additionNumber, substance.inventoryExpenses, substance.finalBalance,
                substance.expenseJustification
            ]
        ws.append(row)

        # Adjusting text alignment and enabling text wrap for better readability
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            ws.row_dimensions[ws.max_row].height = 30  # Set the height for other rows

    
    if month_year == "":
        month_year = datetime.datetime.now().strftime("%B %Y")  # Format: 'Enero 2024'
    default_filename = f"Informe precursores {month_year}.xlsx"
    file_path = asksaveasfilename(defaultextension=".xlsx",
                                filetypes=[("Excel files", "*.xlsx")],
                                initialfile=default_filename,
                                title="Guardar como")
    if file_path:  # Only save if a file path is provided
        wb.save(file_path)
